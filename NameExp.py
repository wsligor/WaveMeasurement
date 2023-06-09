import os, os.path
from typing import Any
from datetime import date
import sqlite3 as sl

import openpyxl as openpyxl
from PySide6.QtCore import Qt, Slot, QModelIndex
from PySide6.QtWidgets import QTextEdit, QLineEdit, QComboBox, QDateEdit, QDialog, QHeaderView, QFileDialog
from PySide6.QtWidgets import QTableView, QMessageBox, QHBoxLayout, QVBoxLayout, QPushButton, QToolButton, QLabel
from PySide6.QtSql import QSqlQueryModel

from Category import dlgCategories
from Group import dlgGroups



class Model(QSqlQueryModel):
    def __init__(self, parent=None):
        super().__init__(parent)

        # self.setHeaderData(0, Qt.Orientation.Horizontal, 'ID')
        self.setHeaderData(1, Qt.Orientation.Horizontal, 'Номер')

        self.refreshNameExp()

    def refreshNameExp(self):
        sql = 'SELECT id, name, date, number FROM nameExp'
        self.setQuery(sql)

    def data(self, item: QModelIndex, role: int = ...) -> Any:
        if not item.isValid():
            return
        if item.column() == 3:
            if role == Qt.ItemDataRole.TextAlignmentRole:
                return Qt.AlignmentFlag.AlignCenter
        return super().data(item, role)

    def addMeanCalcSave(self, values, array):
        con = sl.connect('SFM.db')
        cur = con.cursor()
        sql = '''INSERT INTO nameExp 
                (date, number, name, substance, count, temperature, cuvette, note) 
                    values (?, ?, ?, ?, ?, ?, ?, ?)'''
        cur.execute(sql, values)
        id_nameExp = cur.lastrowid
        self.addDataSetMeanCalcSave(array, id_nameExp, cur)
        con.commit()
        con.close()
        self.refreshNameExp()

    def addDataSetMeanCalcSave(self, array, id_nameExp, cur):
        # book: openpyxl.workbook.workbook.Workbook = openpyxl.load_workbook(filename)
        # sheet: openpyxl.worksheet.worksheet.Worksheet = book.active
        max_rows = 810
        dataExp = []

        for i in range(700):
            waveLength = i + 301
            opticDensity = 0
            transparency = array[i]
            str_dataExp = (waveLength, opticDensity, transparency, id_nameExp)

            dataExp.append(str_dataExp)

        sqlite_insert_query = """INSERT INTO dataExp (waveLength, opticDensity, transparency, id_nameExp)
                               VALUES (?,?,?,?)"""

        cur.executemany(sqlite_insert_query, dataExp)

    def add(self, values, filename):
        con = sl.connect('SFM.db')
        cur = con.cursor()
        sql = '''INSERT INTO nameExp 
                (date, number, name, substance, count, temperature, cuvette, note) 
                    values (?, ?, ?, ?, ?, ?, ?, ?)'''
        cur.execute(sql, values)
        id_nameExp = cur.lastrowid
        self.addDataSet(filename, id_nameExp, cur)
        con.commit()
        con.close()
        self.refreshNameExp()


    def addDataSet(self, filename, id_nameExp, cur):
        book: openpyxl.workbook.workbook.Workbook = openpyxl.load_workbook(filename)
        sheet: openpyxl.worksheet.worksheet.Worksheet = book.active

        max_rows = sheet.max_row
        dataExp = []

        for i in range(2, max_rows + 1):
            waveLength = sheet.cell(row=i, column=1).value
            opticDensity = sheet.cell(row=i, column=2).value
            transparency = sheet.cell(row=i, column=3).value
            str_dataExp = (waveLength, opticDensity, transparency, id_nameExp)

            dataExp.append(str_dataExp)

        sqlite_insert_query = """INSERT INTO dataExp (waveLength, opticDensity, transparency, id_nameExp)
                            VALUES (?,?,?,?)"""

        cur.executemany(sqlite_insert_query, dataExp)


class NameExp(QTableView):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.idSelectTableNameExp = []

        self.model = Model(parent=self)
        self.setModel(self.model)

        # выделение полной строки в таблице
        self.setSelectionBehavior(self.SelectionBehavior.SelectRows)
        # убираем первый столбик ID
        self.hideColumn(0)
        # настройка горизонтального заголовка
        hh = self.horizontalHeader()
        # ширина всех столбцов регулируется по контексту
        hh.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        # столбец "Name" регулируется растягиваясь на оставшуюся длину
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        # настройка вертикального заголовка
        hv = self.verticalHeader()
        # убираем вертикальную нумерацию строк
        hv.hide()

        self.clicked.connect(self.tvNameExp_clicked)

    def tvNameExp_clicked(self):
        i: QModelIndex = self.selectedIndexes()
        # for p in i:
        #     l.append((self.model.index(p.row(), 0).data()))
        # l = [self.model.index(p.row(), 0).data() for p in i]
        # lset = set([self.model.index(p.row(), 0).data() for p in i])
        l = list(set([self.model.index(p.row(), 0).data() for p in i]))
        self.idSelectTableNameExp = l

    @Slot()
    def addNameExp(self):
        values = []
        dia = dlgAddExp()
        if dia.exec():
            values = [dia.dateExp, dia.number, dia.name, dia.substance,
                      dia.countExp, dia.tempExp, dia.cuvette, dia.description]
            self.model.add(values, dia.filename)

    @Slot()
    def addNameExpCalcSave(self, array):
        values = []
        dia = dlgAddExp()
        if dia.exec():
            values = [dia.dateExp, dia.number, dia.name, dia.substance,
                      dia.countExp, dia.tempExp, dia.cuvette, dia.description]
            self.model.addMeanCalcSave(values, array)

    @Slot()
    def updateNameExp(self):
        QMessageBox.information(self, 'NameExp', 'Update')

    @Slot()
    def deleteNameExp(self):
        QMessageBox.information(self, 'NameExp', 'Delete')


class ModelCategories(QSqlQueryModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.refreshCategories()

    def refreshCategories(self):
        sql = 'SELECT name FROM categories'
        self.setQuery(sql)


class ModelGroups(QSqlQueryModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.refreshGroups()

    def refreshGroups(self):
        sql = 'SELECT name FROM groups'
        self.setQuery(sql)


class dlgAddExp(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.__filename = ''

        lblDate = QLabel('Дата', parent=self)
        self.__deDate = QDateEdit(parent=self)
        self.__deDate.setDate(date.today())

        lblNumber = QLabel('Номер', parent=self)
        self.__edNumber = QLineEdit(parent=self)
        self.__edNumber.setText(str(self.returnMaxNumber()))

        lblCategory = QLabel('Категория:', parent=self)
        self.__cbCategory = QComboBox(parent=self)
        self.modelCategories = ModelCategories(parent=self)
        self.__cbCategory.setModel(self.modelCategories)

        lblGroup = QLabel('Группа:', parent=self)
        self.__cbGroup = QComboBox(parent=self)
        self.modelGroups = ModelGroups(parent=self)
        self.__cbGroup.setModel(self.modelGroups)

        lblName = QLabel('Наименование:', parent=self)
        self.__edName = QLineEdit(parent=self)

        lblSubstance = QLabel('Образец:', parent=self)
        self.__edSubstance = QLineEdit(parent=self)

        lblCount = QLabel('% содержания', parent=self)
        self.__edCount = QLineEdit(parent=self)

        lblTemp = QLabel('Температура:', parent=self)
        self.__edTemp = QLineEdit(parent=self)

        lblCuvette = QLabel('Кювета', parent=self)
        self.__edCuvette = QLineEdit(parent=self)

        lblDescription = QLabel('Описание:', parent=self)
        self.__teDescription = QTextEdit(parent=self)

        self.lblFileName = QLabel('Название файла: ', parent=self)
        self.lblFileName.setMinimumHeight(40)
        self.lblFileName.setWordWrap(True)
        self.lblFileName.setAlignment(Qt.AlignmentFlag.AlignTop)

        btnCategory = QToolButton(parent=self)
        btnCategory.setText('...')
        btnGroup = QToolButton(parent=self)
        btnGroup.setText('...')

        btnOk = QPushButton('Ok', parent=self)
        btnCancel = QPushButton('Отмена', parent=self)
        btnFileOpen = QPushButton('Файл', parent=self)

        layHDateNumber = QHBoxLayout()
        layVDate = QVBoxLayout()
        layVDate.addWidget(lblDate)
        layVDate.addWidget(self.__deDate)
        layHDateNumber.addLayout(layVDate)
        layVNumber = QVBoxLayout()
        layVNumber.addWidget(lblNumber)
        layVNumber.addWidget(self.__edNumber)
        layHDateNumber.addLayout(layVNumber)

        layHCategory = QHBoxLayout()
        layHCategory.addWidget(self.__cbCategory)
        layHCategory.addWidget(btnCategory)

        layHCroup = QHBoxLayout()
        layHCroup.addWidget(self.__cbGroup)
        layHCroup.addWidget(btnGroup)

        layHCountTempCuvette = QHBoxLayout()
        layVCount = QVBoxLayout()
        layVCount.addWidget(lblCount)
        layVCount.addWidget(self.__edCount)
        layHCountTempCuvette.addLayout(layVCount)
        layVTemp = QVBoxLayout()
        layVTemp.addWidget(lblTemp)
        layVTemp.addWidget(self.__edTemp)
        layHCountTempCuvette.addLayout(layVTemp)
        layVCuvette = QVBoxLayout()
        layVCuvette.addWidget(lblCuvette)
        layVCuvette.addWidget(self.__edCuvette)
        layHCountTempCuvette.addLayout(layVCuvette)

        layV = QVBoxLayout(self)

        layV.addLayout(layHDateNumber)
        layV.addWidget(lblCategory)
        layV.addLayout(layHCategory)
        layV.addWidget(lblGroup)
        layV.addLayout(layHCroup)
        layV.addWidget(lblName)
        layV.addWidget(self.__edName)
        layV.addWidget(lblSubstance)
        layV.addWidget(self.__edSubstance)
        layV.addLayout(layHCountTempCuvette)
        layV.addWidget(lblDescription)
        layV.addWidget(self.__teDescription)
        layV.addWidget(self.lblFileName)
        layH = QHBoxLayout()
        layH.addWidget(btnOk)
        layH.addWidget(btnCancel)
        layH.addWidget(btnFileOpen)
        layV.addLayout(layH)

        btnCancel.clicked.connect(self.reject)
        btnOk.clicked.connect(self.btnOk_clicked)
        btnFileOpen.clicked.connect(self.btnFileOpenclicked)
        btnCategory.clicked.connect(self.btnCategory_clicked)
        btnGroup.clicked.connect(self.btnGroup_clicked)

    def btnFileOpenclicked(self):
        self.__filename = QFileDialog.getOpenFileName(self, 'Открыть файл', os.getcwd(), 'Excel files (*.xlsx)')[0]
        text = self.lblFileName.text()
        self.lblFileName.setText(text + self.__filename)

    def btnOk_clicked(self):
        if not self.checkForm():
            return
        else:
            self.accept()

    def checkForm(self):
        if self.dateExp is None:
            return False
        if self.number is None:
            return False
        if self.name is None:
            return False
        if self.substance is None:
            return False
        if self.countExp is None:
            return False
        if self.tempExp is None:
            return False
        if self.cuvette is None:
            return False
        if self.filename is None:
            return False
        return True

    def btnGroup_clicked(self):
        dlg_groups = dlgGroups()
        dlg_groups.exec()
        # TODO Установитьь индех по двойному клику
        # row = dlg_groups.tvGroup.currentIndex()
        # p = dlg_groups.tvGroup.model().record(row).value(0)
        # self.__cbGroup.setCurrentIndex(row)
        self.modelGroups.refreshGroups()

    def btnCategory_clicked(self):
        dlg_category = dlgCategories()
        dlg_category.exec()
        self.modelCategories.refreshCategories()

    def returnMaxNumber(self):
        con = sl.connect('SFM.db')
        cur = con.cursor()
        sql = 'SELECT MAX(number) FROM nameExp'
        cur.execute(sql)
        record = cur.fetchone()
        result = record[0] + 1
        con.close()
        return result

    @property
    def filename(self):
        result: str = self.__filename.strip()
        if os.path.exists(self.__filename):
            return result
        else:
            QMessageBox.information(self, 'File', 'Выберите файл')
            return None

    @filename.setter
    def filename(self, value):
        self.__filename = value

    @property
    def number(self):
        result: str = self.__edNumber.text().strip()
        if not result:
            t = 'Номер'
            QMessageBox.information(self, t, 'Заполните поле: ' + t)
            return None
        else:
            return result

    @number.setter
    def number(self, value):
        self.__edNumber.setText((value))

    @property
    def dateExp(self):
        result: str = self.__deDate.text().strip()
        if not result:
            t = 'Дата: '
            QMessageBox.information(self, t, 'Заполните поле: ' + t)
            return None
        else:
            return result

    @dateExp.setter
    def dateExp(self, value):
        self.__deDate.setDate(value)

    @property
    def name(self):
        result: str = self.__edName.text().strip()
        if not result:
            t = 'Наименование'
            QMessageBox.information(self, t, 'Заполните поле: ' + t)
            return None
        else:
            return result

    @name.setter
    def name(self, value):
        self.__edName.setText(value)

    @property
    def substance(self):
        result: str = self.__edSubstance.text().strip()
        if not result:
            t = 'Образец'
            QMessageBox.information(self, t, 'Заполните поле: ' + t)
            return None
        else:
            return result

    @substance.setter
    def substance(self, value):
        self.__edSubstance.setText(value)

    @property
    def countExp(self):
        result: str = self.__edCount.text().strip()
        if not result:
            t = 'Процент'
            QMessageBox.information(self, t, 'Заполните поле: ' + t)
            return None
        else:
            return result

    @countExp.setter
    def countExp(self, value):
        self.__edCount.setText(value)

    @property
    def tempExp(self):
        result: str = self.__edTemp.text().strip()
        if not result:
            t = 'Температура'
            QMessageBox.information(self, t, 'Заполните поле: ' + t)
            return None
        else:
            return result

    @tempExp.setter
    def tempExp(self, value):
        self.__edTemp.setText(value)

    @property
    def cuvette(self):
        result: str = self.__edCuvette.text().strip()
        if not result:
            t = 'Кювета'
            QMessageBox.information(self, t, 'Заполните поле: ' + t)
            return None
        else:
            return result

    @cuvette.setter
    def cuvette(self, value):
        self.__edCuvette.setText(value)

    @property
    def description(self):
        result: str = self.__teDescription.toPlainText().strip()
        if not result:
            return None
        else:
            return result

    @description.setter
    def description(self, value):
        self.__teDescription.setPlainText(value)


